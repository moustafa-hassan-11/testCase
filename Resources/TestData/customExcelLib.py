from _csv import reader
import openpyxl

from pathlib import Path

from robot.libraries.BuiltIn import BuiltIn


def GetVariablesFromExcel(sheetName):

    xlsx_file = Path('./Resources/TestData/', 'TA Catalog.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.get_sheet_by_name(sheetName)
    tee = []
    for i in range(sheet.max_column):

        for j in (1, sheet.max_row):
            tee.append(sheet[j][i].value)
        BuiltIn().set_suite_variable('${' + tee[0] + '}', tee[1])
        #print(tee[0], tee[1])
        tee.clear()


